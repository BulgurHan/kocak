from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render,redirect
from django.contrib.auth import login,authenticate,logout
from openpyxl import Workbook,load_workbook
from .forms import BordroForm, GirisForm,MukellefForm, OzlukForm,SubeForm,PersonelForm,AramaForm, ToDoForm,TutanakForm
from .models import Mukellef, Ozluk,Sube,Personel,Bordro,Tutanak,ToDo



def signin(request):
    if request.user.is_authenticated:
        return redirect("main")   
    if request.method == "POST":
        form = GirisForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request,username=username,password=password)
            if user is not None:
                if user.is_active:
                    login(request,user)
                    return redirect("main")
            else:
                messages.info(request,"Kullanıcı Adı veya Şifre Yanlış")
    else:
        form= GirisForm()
    return render(request,"signin.html",{"form":form})

@login_required(login_url="/")
def main(request):
    context=dict()
    ozlukler = Ozluk.objects.all()
    bordrolar = Bordro.objects.all()
    context['bordrolar'] = list()
    context['items'] = list()
    context["toplam"] = len(context["items"])
    context["bordrolar_toplam"] = len(context["bordrolar"])
    context['todos'] = ToDo.objects.all()
    context['form'] = ToDoForm(request.POST)
    for ozluk in ozlukler:  
        if ozluk.is_sozlesmesi == False or ozluk.gorevlendirme_belgesi == False or ozluk.fazla_messai_onay == False or ozluk.gecici_gorev_yazisi == False or ozluk.kimlik_fotokopisi == False or ozluk.adli_sicil_kaydi == False or ozluk.yerlesim_yeri_belgesi == False or ozluk.mezuniyet_belgesi == False or ozluk.gorevlendirme_belgesi == False or ozluk.sertifikalar == False or ozluk.hijyen_belgesi == False or ozluk.saglik_raporu == False :
            context['items'].append(ozluk)
        if ozluk.istifa == False:
            continue
        else:
            ozluk.personel.tamamlandi = True
            ozluk.personel.save()
    if len(context["items"]) > 0:
        messages.info(request,"{} Personelin Özlük Dosyasında Eksik Var!".format(len(context["items"])))
    for bordro in bordrolar:
        if bordro.geldi == False:
            context["bordrolar"].append(bordro)
    if len(context['bordrolar']) > 0 :
        messages.info(request,"{} Personelin Bordrosu Girilmiş ama Gelmemiş!".format(len(context["bordrolar"])))
    if request.method == "POST":
        if context['form'].is_valid():
            aciklama = context['form'].cleaned_data['aciklama']
            ToDo.objects.create(
                aciklama = aciklama,
                ok = False
            )
    return render(request,"main.html",context)

def todo_degis(request,todo_pk):
    todo = ToDo.objects.get(pk=todo_pk)
    todo.ok = True
    todo.save()
    return redirect("main")

def todo_sil(request,todo_pk):
    todo = ToDo.objects.get(pk=todo_pk)
    todo.delete() 
    return redirect("main")

@login_required(login_url="/")
def bordrolar(request):
    context = dict()
    context['title'] = "Bordrosu Girilen fakat Gelmeyen Personeller"
    bordrolar = Bordro.objects.all()
    context['bordrolar'] = list()
    for bordro in bordrolar:
        if bordro.geldi == False:
            context['bordrolar'].append(bordro)
    return render(request,"bordrolar.html",context)


@login_required(login_url="/")
def mukellefler(request):
    context = dict()
    context['mukellefler'] = Mukellef.objects.all()
    context['subeler'] = Sube.objects.all()
    return render(request,"mukellefler.html",context)



@login_required(login_url="/")
def personeller(request):
    context = dict()
    context['title'] = "Personeller"
    context["items"] = Personel.objects.all()
    context["mukellefler"] = Mukellef.objects.all()
    context['form'] = AramaForm(request.POST)
    if request.method=="POST":
        context['form'] = AramaForm(request.POST)
        if context['form'].is_valid():
            kelime = context['form'].cleaned_data['kelime']
            return redirect("sonuc", kelime=kelime)
    return render(request,"personeller.html",context)


@login_required(login_url="/")
def bordro_ekle(request,personel_pk):
    context = dict()
    personel = Personel.objects.get(pk=personel_pk)
    context['form'] = BordroForm(request.POST)
    if request.method == "POST":
        if context['form'].is_valid():
            tarih = context['form'].cleaned_data['tarih']
            geldi = context['form'].cleaned_data['geldi']
            Bordro.objects.create(
                personel=personel,
                tarih=tarih,
                geldi=geldi
            )
            return redirect("personel", personel_pk=personel_pk)
    return render(request,'form_base.html',context)

@login_required(login_url="/")
def tutanak_ekle(request,personel_pk):
    context = dict()
    personel = Personel.objects.get(pk=personel_pk)
    context['form'] = TutanakForm(request.POST)
    if request.method == "POST":
        if context['form'].is_valid():
            baslik = context['form'].cleaned_data['baslik']
            tarih = context['form'].cleaned_data['tarih']
            Tutanak.objects.create(
                personel=personel,
                tarih=tarih,
                baslik=baslik
            )
            return redirect("personel", personel_pk=personel_pk)
    return render(request,'form_base.html',context)

@login_required(login_url="/")
def tutanak_sil(request,tutanak_pk,personel_pk):
    tutanak = Tutanak.objects.get(pk=tutanak_pk)
    tutanak.delete()
    return redirect("personel", personel_pk=personel_pk)

@login_required(login_url="/")
def bordro_geldi(request,bordro_pk,personel_pk):
    bordro = Bordro.objects.get(pk=bordro_pk)
    bordro.geldi = True
    bordro.save()
    return redirect("personel", personel_pk=personel_pk)

@login_required(login_url="/")
def bordro_sil(request,bordro_pk,personel_pk):
    bordro = Bordro.objects.get(pk=bordro_pk)
    bordro.delete()
    return redirect("personel", personel_pk=personel_pk)


@login_required(login_url="/")
def eksikler(request):
    context = dict()
    context['title'] = "Eksiği Olan Personeller"
    context["items"] = Personel.objects.filter(tamamlandi=False)
    context["mukellefler"] = Mukellef.objects.all()
    context['form'] = AramaForm(request.POST)
    if request.method=="POST":
        context['form'] = AramaForm(request.POST)
        if context['form'].is_valid():
            kelime = context['form'].cleaned_data['kelime']
            return redirect("sonuc", kelime=kelime)
    return render(request,"personeller.html",context)



@login_required(login_url="/")
def sonuc(request,kelime):
    context = dict()
    context["items"] = Personel.objects.filter(isim=kelime)
    return render(request,"sonuc.html",context)

@login_required(login_url="/")
def personel(request,personel_pk):
    context = dict()
    personel = Personel.objects.get(pk=personel_pk)
    ozluk = Ozluk.objects.get(personel=personel)
    context['form'] = OzlukForm(data=request.POST, files = request.FILES,instance=ozluk)   
    context["personel"] = Personel.objects.get(pk=personel_pk)
    context["ozluk" ]= Ozluk.objects.get(personel=personel)
    context['bordrolar'] = Bordro.objects.filter(personel=personel)
    context["tutanaklar"] = Tutanak.objects.filter(personel=personel)
    if request.method == "POST":
        if context['form'].is_valid():
            context['form'].save()
            return redirect("personel", personel_pk=personel_pk)
    return render(request,"personel.html",context)

@login_required(login_url="/")
def tanimlamalar(request):
    context = dict()
    return render(request,"tanimlamalar.html",context)

@login_required(login_url="/")
def mukellef_tanımlama(request):
    context = dict()
    context['title'] = "Mükellef"
    context['items'] = Mukellef.objects.all()
    context['toplam'] = len(Mukellef.objects.all())
    return render(request,"tanımlama_base.html",context)
@login_required(login_url="/")
def mukellef_ekle(request):
    context=dict()
    context['form'] = MukellefForm(request.POST)
    if request.method =="POST":      
        if context['form'].is_valid():
            context['form'].save()
            return redirect("mukellef_tanımlama")
    return render(request,"form.html",context)

@login_required(login_url="/")
def sube_tanımlama(request):
    context = dict()
    context['title'] = "Şube"
    context['form'] = SubeForm(request.POST)
    context['items'] = Sube.objects.all().order_by("mukellef")
    context['toplam'] = len(Sube.objects.all())
    if request.method == "POST":
        if context['form'].is_valid():
            context['form'].save()   
    return render(request,"tanımlama_base.html",context)

@login_required(login_url="/")
def toplu_mukellef(request):
    context = dict()
    context['title'] = "Mükellef"
    if request.method == "POST":
        data = request.FILES.get("data")
        if data:
            wb = load_workbook(data)
            ws = wb.active
            mailler = list()
            isimler = list()
            adresler = list()
            teller = list()
            for satir in range(1,ws.max_row+1):
                if str(ws.cell(satir,1).value) != "İsim":
                    isimler.append(str(ws.cell(satir,1).value).strip())
                if str(ws.cell(satir,2).value) != "Adres":
                    adresler.append(str(ws.cell(satir,2).value).strip())
                if str(ws.cell(satir,3).value) != "Tel":
                    teller.append(str(ws.cell(satir,3).value).strip())
                if str(ws.cell(satir,4).value) != "Mail":
                    mailler.append(str(ws.cell(satir,4).value).strip())
            
            i = 0
            while i < len(isimler):
                Mukellef.objects.create(
                    isim = isimler[i],
                    adres = adresler[i],
                    tel = teller[i],
                    mail = mailler[i]
                )
                i+=1                                      
            return redirect("mukellef_tanımlama")                           
                                                                    
    return render(request,"toplu_mukellef.html",context)

@login_required(login_url="/")
def toplu_sube(request):
    context = dict()
    context['title'] = "Şube"
    if request.method == "POST":
        data = request.FILES.get("data")
        if data:
            wb = load_workbook(data)
            ws = wb.active
            isimler = list()
            mukellefler = list()
            mukellef_items = list()
            for satir in range(1,ws.max_row+1):
                if str(ws.cell(satir,1).value) != "Mükellef":
                    mukellefler.append(str(ws.cell(satir,1).value).strip())
                if str(ws.cell(satir,2).value) != "İsim":
                    isimler.append(str(ws.cell(satir,2).value).strip())
            for mukellef in mukellefler:
                item = Mukellef.objects.get(isim=mukellef)
                mukellef_items.append(item)
            i = 0
            while i< len(mukellefler):
                Sube.objects.create(
                    isim = isimler[i],
                    mukellef = mukellef_items[i]
                )
                i += 1
            return redirect("sube_tanımlama")                                                
    return render(request,"toplu_mukellef.html",context)

@login_required(login_url="/")
def personel_tanımlama(request):
    context = dict()
    context['title'] = "Personel"
    context['items'] = Personel.objects.all()
    context['mukellefler'] = Mukellef.objects.all()
    context['toplam'] = len(Personel.objects.all())
    return render(request,"tanımlama_base.html",context)

@login_required(login_url="/")
def personel_ekle(request):
    context=dict()
    context['form'] = PersonelForm(request.POST)
    if request.method =="POST":      
        if context['form'].is_valid():
            context['form'].save()
            return redirect("personel_tanımlama")
    return render(request,"personel_form.html",context)


@login_required(login_url="/")
def toplu_personel(request):
    context = dict()
    context['title'] = "Personel"
    if request.method == "POST":
        data = request.FILES.get("data")
        if data:
            wb = load_workbook(data)
            ws = wb.active
            mukellefler = list()
            subeler = list()
            isimler = list()
            tcler = list()
            adresler = list()
            meslekler = list()
            ise_girisler = list()
            personeller=list()
            for satir in range(1,ws.max_row+1):
                if str(ws.cell(satir,1).value) != "Mükellef":
                    a = str(ws.cell(satir,1).value).strip()
                    mukellef = Mukellef.objects.get(isim=a)
                    mukellefler.append(mukellef)
                if str(ws.cell(satir,2).value) != "Şube":
                    try:
                        c = str(ws.cell(satir,2).value).strip()
                        sube = Sube.objects.get(mukellef=mukellef,isim=c)
                        subeler.append(sube)
                    except:
                        messages.warning(request,"{} in {} adında bir şubesi yok".format(mukellef,c))
                        return redirect("personel_tanımlama")
                if str(ws.cell(satir,3).value) != "İsim Soyisim":
                    isimler.append(str(ws.cell(satir,3).value))
                if str(ws.cell(satir,4).value) != "T.C":
                    tcler.append(str(ws.cell(satir,4).value))
                if str(ws.cell(satir,5).value) != "Adres":
                    adresler.append(str(ws.cell(satir,5).value))                        
                if str(ws.cell(satir,6).value) != "Meslek":
                    meslekler.append(str(ws.cell(satir,6).value))
                if str(ws.cell(satir,7).value) != "İşe Giriş":
                    ise_girisler.append(str(ws.cell(satir,7).value))
            i = 0
            while i<len(mukellefler):
                a = Personel.objects.create(
                    mukellef = mukellefler[i],
                    sube = subeler[i],
                    isim = isimler[i],
                    tc = tcler[i],
                    adres = adresler[i],
                    meslek=meslekler[i]
                )
                personeller.append(a)
                i +=1
            for personel in personeller:
                Ozluk.objects.create(
                    personel=personel
                )                
            return redirect("personel_tanımlama")                           
                                                                    
    return render(request,"toplu_mukellef.html",context)



    